# bott.py
import asyncio
import sqlite3
import os
import time
import jdatetime
from openpyxl import Workbook

from aiogram import Bot, Dispatcher, types
from aiogram.types import (
    KeyboardButton,
    ReplyKeyboardMarkup,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
)

# ------------------------
# CONFIG
# ------------------------
#TOKEN = "8396998653:AAGLdV013TRrwVXfLepgspAU7mIBW9bsliU"  # <-- توکن ربات را اینجا قرار بده
import os
TOKEN = os.getenv("BOT_TOKEN")
DB_PATH = "database.db"
EXCEL_PATH = "report.xlsx"

bot = Bot(token=TOKEN)
dp = Dispatcher()

# ------------------------
# DATABASE (SQLite)
# ------------------------
conn = sqlite3.connect(DB_PATH, check_same_thread=False)
cursor = conn.cursor()
cursor.execute("""
CREATE TABLE IF NOT EXISTS employees(
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL
)
""")
cursor.execute("""
CREATE TABLE IF NOT EXISTS projects(
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL
)
""")
cursor.execute("""
CREATE TABLE IF NOT EXISTS work_logs(
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    employee_id INTEGER,
    project_id INTEGER,
    date TEXT,        -- شمسی YYYY-MM-DD
    hours REAL,
    overtime REAL
)
""")
conn.commit()

# ------------------------
# BUTTON TEXTS
# ------------------------
BTN_MANAGE_EMP = "👤 مدیریت کارمندان"
BTN_MANAGE_PROJ = "🏗 مدیریت پروژه‌ها"
BTN_REGISTER = "🕒 ثبت ساعت تکی"
BTN_GROUP_REGISTER = "👥 ثبت گروهی ساعت"
BTN_REPORT = "📊 گزارش تفکیکی"
BTN_EXPORT = "📥 خروجی اکسل"
BTN_BACK = "⬅ بازگشت"
BTN_ADD_EMP = "➕ افزودن کارمند"
BTN_LIST_EMP = "📋 لیست کارمندان"
BTN_DEL_EMP = "🗑 حذف کارمند"
BTN_ADD_PROJ = "➕ افزودن پروژه"
BTN_LIST_PROJ = "📋 لیست پروژه‌ها"
BTN_DEL_PROJ = "🗑 حذف پروژه"
BTN_CONTINUE = "✅ ادامه ثبت"
BTN_FINISH = "🏁 اتمام ثبت"

# ------------------------
# KEYBOARDS
# ------------------------
def main_menu():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=BTN_MANAGE_EMP), KeyboardButton(text=BTN_MANAGE_PROJ)],
            [KeyboardButton(text=BTN_REGISTER), KeyboardButton(text=BTN_GROUP_REGISTER)],
            [KeyboardButton(text=BTN_REPORT), KeyboardButton(text=BTN_EXPORT)],
        ],
        resize_keyboard=True,
    )

def back_kb():
    return ReplyKeyboardMarkup(keyboard=[[KeyboardButton(text=BTN_BACK)]], resize_keyboard=True, one_time_keyboard=True)

def continue_kb():
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text=BTN_CONTINUE), KeyboardButton(text=BTN_FINISH)]],
        resize_keyboard=True,
        one_time_keyboard=True,
    )

def employees_kb():
    rows = cursor.execute("SELECT id, name FROM employees ORDER BY id").fetchall()
    if not rows:
        return ReplyKeyboardMarkup(keyboard=[[KeyboardButton(text=BTN_BACK)]], resize_keyboard=True)
    kb = [[KeyboardButton(text=f"{r[0]} - {r[1]}")] for r in rows]
    kb.append([KeyboardButton(text=BTN_BACK)])
    return ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True, one_time_keyboard=True)

def projects_kb():
    rows = cursor.execute("SELECT id, name FROM projects ORDER BY id").fetchall()
    if not rows:
        return ReplyKeyboardMarkup(keyboard=[[KeyboardButton(text=BTN_BACK)]], resize_keyboard=True)
    kb = [[KeyboardButton(text=f"{r[0]} - {r[1]}")] for r in rows]
    kb.append([KeyboardButton(text=BTN_BACK)])
    return ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True, one_time_keyboard=True)

# ------------------------
# INLINE helpers
# ------------------------
def inline_employee_multi(selected_ids=None):
    if selected_ids is None:
        selected_ids = []
    rows = cursor.execute("SELECT id, name FROM employees ORDER BY id").fetchall()
    kb = []
    for emp_id, name in rows:
        mark = "✔" if emp_id in selected_ids else "✖"
        kb.append([InlineKeyboardButton(text=f"{mark} {name}", callback_data=f"empgrp_{emp_id}")])
    kb.append([
        InlineKeyboardButton(text="✔ تأیید", callback_data="empgrp_confirm"),
        InlineKeyboardButton(text="❌ انصراف", callback_data="empgrp_cancel"),
    ])
    return InlineKeyboardMarkup(inline_keyboard=kb)

def projects_inline():
    rows = cursor.execute("SELECT id, name FROM projects ORDER BY id").fetchall()
    kb = []
    for pid, name in rows:
        kb.append([InlineKeyboardButton(text=name, callback_data=f"grpprj_{pid}")])
    kb.append([InlineKeyboardButton(text="❌ انصراف", callback_data="grpprj_cancel")])
    return InlineKeyboardMarkup(inline_keyboard=kb)

# ------------------------
# CALENDAR (Jalali simple)
# ------------------------
def jdate_to_str(jdate: jdatetime.date) -> str:
    return f"{jdate.year:04d}-{jdate.month:02d}-{jdate.day:02d}"

def calendar_inline(jdate: jdatetime.date):
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=f"📅 {jdate_to_str(jdate)}", callback_data="noop")],
        [
            InlineKeyboardButton(text="⬅️ یک روز قبل", callback_data="cal_prev"),
            InlineKeyboardButton(text="روز بعد ➡️", callback_data="cal_next")
        ],
        [
            InlineKeyboardButton(text="امروز", callback_data="cal_today"),
            InlineKeyboardButton(text="دیروز", callback_data="cal_yesterday")
        ],
        [
            InlineKeyboardButton(text="✔️ تأیید تاریخ", callback_data="cal_confirm"),
            InlineKeyboardButton(text="❌ انصراف", callback_data="cal_cancel")
        ]
    ])

# ------------------------
# STATE (simple FSM)
# ------------------------
user_sessions = {}  # uid -> {"state": str, "data": dict}

def set_state(uid, state, data=None):
    user_sessions[uid] = {"state": state, "data": (data.copy() if isinstance(data, dict) else {}) if data else {}}

def get_state(uid):
    return user_sessions.get(uid, {"state": None, "data": {}})

def clear_state(uid):
    if uid in user_sessions:
        del user_sessions[uid]

# ------------------------
# UTIL
# ------------------------
def parse_id_name(text: str):
    if not text or "-" not in text:
        return None, None
    parts = text.split(" - ", 1)
    try:
        _id = int(parts[0].strip())
    except:
        return None, None
    name = parts[1].strip() if len(parts) > 1 else ""
    return _id, name

# ------------------------
# MAIN MESSAGE HANDLER
# ------------------------
@dp.message()
async def handler(msg: types.Message):
    text = (msg.text or "").strip()
    uid = msg.from_user.id

    # ensure session
    if uid not in user_sessions:
        user_sessions[uid] = {"state": None, "data": {}}

    st = get_state(uid)["state"]
    data = get_state(uid)["data"]

    # PRIORITY: handle deletion states first
    if st == "delete_employee_select":
        emp_id, _ = parse_id_name(text)
        if not emp_id:
            await msg.answer("لطفاً از دکمه‌ها انتخاب کن.", reply_markup=employees_kb())
            return
        cursor.execute("DELETE FROM employees WHERE id=?", (emp_id,))
        conn.commit()
        clear_state(uid)
        await msg.answer("✔ کارمند حذف شد.", reply_markup=main_menu())
        return

    if st == "delete_project_select":
        proj_id, _ = parse_id_name(text)
        if not proj_id:
            await msg.answer("لطفاً از دکمه‌ها انتخاب کن.", reply_markup=projects_kb())
            return
        cursor.execute("DELETE FROM projects WHERE id=?", (proj_id,))
        conn.commit()
        clear_state(uid)
        await msg.answer("✔ پروژه حذف شد.", reply_markup=main_menu())
        return

    # Buttons that bypass state
    if text == BTN_CONTINUE:
        clear_state(uid)
        set_state(uid, "await_employee_select", {})
        await msg.answer("ادامه ثبت — یک کارمند انتخاب کنید:", reply_markup=employees_kb())
        return

    if text == BTN_FINISH:
        clear_state(uid)
        await msg.answer("ثبت ساعت کاری پایان یافت.", reply_markup=main_menu())
        return

    if text == "/start":
        clear_state(uid)
        await msg.answer("سلام! به ربات مدیریت پروژه خوش آمدی.", reply_markup=main_menu())
        return

    if text == BTN_BACK:
        clear_state(uid)
        await msg.answer("بازگشت به منوی اصلی.", reply_markup=main_menu())
        return

    # refresh state
    st = get_state(uid)["state"]
    data = get_state(uid)["data"]

    # If in a state -> handle
    if st not in (None, ""):
        # add employee
        if st == "await_employee_name":
            name = text.strip()
            if not name:
                await msg.answer("نام معتبر نیست. دوباره تلاش کن.", reply_markup=back_kb())
                return
            cursor.execute("INSERT INTO employees(name) VALUES(?)", (name,))
            conn.commit()
            clear_state(uid)
            await msg.answer(f"✔ کارمند «{name}» اضافه شد.", reply_markup=main_menu())
            return

        # add project
        if st == "await_project_name":
            name = text.strip()
            if not name:
                await msg.answer("نام معتبر نیست. دوباره تلاش کن.", reply_markup=back_kb())
                return
            cursor.execute("INSERT INTO projects(name) VALUES(?)", (name,))
            conn.commit()
            clear_state(uid)
            await msg.answer(f"✔ پروژه «{name}» اضافه شد.", reply_markup=main_menu())
            return

        # report by employee select
        if st == "report_employee_select":
            emp_id, emp_name = parse_id_name(text)
            if not emp_id:
                await msg.answer("لطفا از دکمه‌ها استفاده کنید.", reply_markup=employees_kb())
                return
            rows = cursor.execute("""
                SELECT p.name, w.date, w.hours, w.overtime
                FROM work_logs w
                JOIN projects p ON p.id = w.project_id
                WHERE w.employee_id = ?
                ORDER BY w.date
            """, (emp_id,)).fetchall()
            if not rows:
                clear_state(uid)
                await msg.answer("هیچ رکوردی برای این کارمند وجود ندارد.", reply_markup=main_menu())
                return
            wb = Workbook()
            ws = wb.active
            ws.title = "employee_report"
            ws.append(["project", "date", "hours", "overtime"])
            for r in rows:
                ws.append([r[0], r[1], r[2], r[3]])
            safe = (emp_name or f"emp_{emp_id}").replace("/", "-").replace("\\", "-")
            filename = f"emp-report-{safe}.xlsx"
            wb.save(filename)
            await msg.answer("در حال ارسال فایل گزارش...")
            await bot.send_document(msg.chat.id, types.FSInputFile(filename))
            try: os.remove(filename)
            except: pass
            clear_state(uid)
            await msg.answer("انجام شد.", reply_markup=main_menu())
            return

        # report by project select
        if st == "report_project_select":
            proj_id, proj_name = parse_id_name(text)
            if not proj_id:
                await msg.answer("لطفا از دکمه‌ها استفاده کنید.", reply_markup=projects_kb())
                return
            rows = cursor.execute("""
                SELECT e.name, w.date, w.hours, w.overtime
                FROM work_logs w
                JOIN employees e ON e.id = w.employee_id
                WHERE w.project_id = ?
                ORDER BY w.date
            """, (proj_id,)).fetchall()
            if not rows:
                clear_state(uid)
                await msg.answer("هیچ رکوردی برای این پروژه وجود ندارد.", reply_markup=main_menu())
                return
            wb = Workbook()
            ws = wb.active
            ws.title = "project_report"
            ws.append(["employee", "date", "hours", "overtime"])
            for r in rows:
                ws.append([r[0], r[1], r[2], r[3]])
            safe = (proj_name or f"proj_{proj_id}").replace("/", "-").replace("\\", "-")
            filename = f"proj-report-{safe}.xlsx"
            wb.save(filename)
            await msg.answer("در حال ارسال فایل گزارش...")
            await bot.send_document(msg.chat.id, types.FSInputFile(filename))
            try: os.remove(filename)
            except: pass
            clear_state(uid)
            await msg.answer("انجام شد.", reply_markup=main_menu())
            return

        # register flow states (t e n s)
        if st == "await_employee_select":
            emp_id, _ = parse_id_name(text)
            if not emp_id:
                await msg.answer("لطفاً از دکمه‌ها انتخاب کن.", reply_markup=employees_kb())
                return
            set_state(uid, "await_project_select", {"emp_id": emp_id})
            await msg.answer("پروژه را انتخاب کنید:", reply_markup=projects_kb())
            return

        if st == "await_project_select":
            proj_id, _ = parse_id_name(text)
            if not proj_id:
                await msg.answer("لطفاً از دکمه‌ها انتخاب کن.", reply_markup=projects_kb())
                return
            jd = jdatetime.date.today()
            newdata = {"emp_id": data["emp_id"], "proj_id": proj_id, "jdate": jd}
            set_state(uid, "await_calendar", newdata)
            sent = await msg.answer("تاریخ را انتخاب کنید:", reply_markup=calendar_inline(jd))
            user_sessions[uid]["calendar_message_id"] = sent.message_id
            return

        if st == "await_hours":
            try:
                hours = float(text)
            except:
                await msg.answer("لطفا عدد صحیح یا اعشاری وارد کن (مثلا 8 یا 7.5).", reply_markup=back_kb())
                return
            data["hours"] = hours
            set_state(uid, "await_overtime", data)
            await msg.answer("اضافه‌کاری را وارد کنید (مثلا 0 یا 1.5):", reply_markup=back_kb())
            return

        if st == "await_overtime":
            try:
                ot = float(text)
            except:
                await msg.answer("لطفا عدد صحیح یا اعشاری وارد کن (مثلا 0 یا 1.5).", reply_markup=back_kb())
                return
            if not all(k in data for k in ("emp_id", "proj_id", "date_jsh", "hours")):
                clear_state(uid)
                await msg.answer("خطا در داده‌ها. لطفاً دوباره ثبت را شروع کن.", reply_markup=main_menu())
                return
            cursor.execute("""
                INSERT INTO work_logs(employee_id, project_id, date, hours, overtime)
                VALUES (?, ?, ?, ?, ?)
            """, (data["emp_id"], data["proj_id"], data["date_jsh"], data["hours"], ot))
            conn.commit()
            set_state(uid, "after_saved", {})
            await msg.answer("✔ رکورد ذخیره شد.", reply_markup=continue_kb())
            return

        # --- GROUP FLOW STATES ---
        if st == "group_hours":
            try:
                hours = float(text)
            except:
                await msg.answer("⛔ عدد نامعتبر — دوباره وارد کن:", reply_markup=back_kb())
                return
            employees = data.get("employees", [])
            data["hours"] = hours
            set_state(uid, "group_overtime", data)
            await msg.answer("اضافه کاری مشترک را وارد کنید (مثلا 0 یا 1.5):", reply_markup=back_kb())
            return

        if st == "group_overtime":
            try:
                ot = float(text)
            except:
                await msg.answer("⛔ عدد نامعتبر — دوباره وارد کن:", reply_markup=back_kb())
                return
            # final validation
            if not all(k in data for k in ("employees", "project_id", "date_jsh", "hours")):
                clear_state(uid)
                await msg.answer("خطا در داده‌ها. لطفاً دوباره ثبت گروهی را شروع کن.", reply_markup=main_menu())
                return
            employees = data["employees"]
            hours = data["hours"]
            date_sh = data["date_jsh"]
            proj_id = data["project_id"]
            for emp_id in employees:
                cursor.execute("""
                    INSERT INTO work_logs(employee_id, project_id, date, hours, overtime)
                    VALUES (?, ?, ?, ?, ?)
                """, (emp_id, proj_id, date_sh, hours, ot))
            conn.commit()
            clear_state(uid)
            await msg.answer(f"✔ ثبت گروهی برای {len(employees)} کارمند انجام شد.", reply_markup=main_menu())
            return

        # unknown state fallback
        await msg.answer("شما در یک حالت داخلی هستی. از منوی اصلی استفاده کن یا بازگشت بزن.", reply_markup=main_menu())
        clear_state(uid)
        return

    # ------------------------
    # NOT IN STATE -> normal menu handling
    # ------------------------
    if text == BTN_MANAGE_EMP:
        clear_state(uid)
        kb = ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text=BTN_ADD_EMP)],
                [KeyboardButton(text=BTN_LIST_EMP)],
                [KeyboardButton(text=BTN_DEL_EMP)],
                [KeyboardButton(text=BTN_BACK)]
            ],
            resize_keyboard=True,
        )
        await msg.answer("مدیریت کارمندان:", reply_markup=kb)
        return

    if text == BTN_ADD_EMP:
        set_state(uid, "await_employee_name", {})
        await msg.answer("نام کارمند را وارد کنید:", reply_markup=back_kb())
        return

    if text == BTN_LIST_EMP:
        rows = cursor.execute("SELECT id, name FROM employees ORDER BY id").fetchall()
        if not rows:
            await msg.answer("هیچ کارمندی ثبت نشده.", reply_markup=main_menu())
        else:
            s = "لیست کارمندان:\n\n" + "\n".join(f"{r[0]} - {r[1]}" for r in rows)
            await msg.answer(s, reply_markup=main_menu())
        return

    if text == BTN_DEL_EMP:
        rows = cursor.execute("SELECT id, name FROM employees ORDER BY id").fetchall()
        if not rows:
            await msg.answer("هیچ کارمندی برای حذف وجود ندارد.", reply_markup=main_menu())
            return
        kb = [[KeyboardButton(text=f"{r[0]} - {r[1]}")] for r in rows]
        kb.append([KeyboardButton(text=BTN_BACK)])
        await msg.answer("کدام کارمند حذف شود؟", reply_markup=ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True, one_time_keyboard=True))
        set_state(uid, "delete_employee_select", {})
        return

    # PROJECTS
    if text == BTN_MANAGE_PROJ:
        clear_state(uid)
        kb = ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text=BTN_ADD_PROJ)],
                [KeyboardButton(text=BTN_LIST_PROJ)],
                [KeyboardButton(text=BTN_DEL_PROJ)],
                [KeyboardButton(text=BTN_BACK)]
            ],
            resize_keyboard=True
        )
        await msg.answer("مدیریت پروژه‌ها:", reply_markup=kb)
        return

    if text == BTN_ADD_PROJ:
        set_state(uid, "await_project_name", {})
        await msg.answer("نام پروژه را وارد کنید:", reply_markup=back_kb())
        return

    if text == BTN_LIST_PROJ:
        rows = cursor.execute("SELECT id, name FROM projects ORDER BY id").fetchall()
        if not rows:
            await msg.answer("هیچ پروژه‌ای ثبت نشده.", reply_markup=main_menu())
        else:
            s = "لیست پروژه‌ها:\n\n" + "\n".join(f"{r[0]} - {r[1]}" for r in rows)
            await msg.answer(s, reply_markup=main_menu())
        return

    if text == BTN_DEL_PROJ:
        rows = cursor.execute("SELECT id, name FROM projects ORDER BY id").fetchall()
        if not rows:
            await msg.answer("هیچ پروژه‌ای برای حذف وجود ندارد.", reply_markup=main_menu())
            return
        kb = [[KeyboardButton(text=f"{r[0]} - {r[1]}")] for r in rows]
        kb.append([KeyboardButton(text=BTN_BACK)])
        await msg.answer("کدام پروژه حذف شود؟", reply_markup=ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True, one_time_keyboard=True))
        set_state(uid, "delete_project_select", {})
        return

    # REPORTS
    if text == BTN_REPORT:
        clear_state(uid)
        kb = ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="📄 گزارش بر اساس کارمند")],
                [KeyboardButton(text="📄 گزارش بر اساس پروژه")],
                [KeyboardButton(text=BTN_BACK)]
            ],
            resize_keyboard=True
        )
        await msg.answer("نوع گزارش را انتخاب کنید:", reply_markup=kb)
        return

    if text == "📄 گزارش بر اساس کارمند":
        set_state(uid, "report_employee_select", {})
        await msg.answer("کارمند را انتخاب کنید:", reply_markup=employees_kb())
        return

    if text == "📄 گزارش بر اساس پروژه":
        set_state(uid, "report_project_select", {})
        await msg.answer("پروژه را انتخاب کنید:", reply_markup=projects_kb())
        return

    # EXPORT ALL
    if text == BTN_EXPORT:
        await generate_and_send_excel(msg)
        return

    # REGISTER (t e k i)
    if text == BTN_REGISTER:
        rows = cursor.execute("SELECT id, name FROM employees ORDER BY id").fetchall()
        if not rows:
            await msg.answer("هیچ کارمندی ثبت نشده.", reply_markup=main_menu())
            return
        set_state(uid, "await_employee_select", {})
        await msg.answer("یک کارمند انتخاب کنید:", reply_markup=employees_kb())
        return

    # GROUP REGISTER
    if text == BTN_GROUP_REGISTER:
        rows = cursor.execute("SELECT id, name FROM employees ORDER BY id").fetchall()
        if not rows:
            await msg.answer("هیچ کارمندی ثبت نشده.", reply_markup=main_menu())
            return
        set_state(uid, "group_select_employees", {"selected": []})
        await msg.answer("کارمندان را انتخاب کنید (✔ یعنی انتخاب شده):", reply_markup=inline_employee_multi([]))
        return

    # fallback
    await msg.answer("متوجه نشدم. لطفاً از منوی اصلی استفاده کن.", reply_markup=main_menu())
    clear_state(uid)
    return

# ------------------------
# CALLBACK QUERY HANDLER (calendar + group selection + project choose)
# ------------------------
@dp.callback_query()
async def cal_cb(cq: types.CallbackQuery):
    uid = cq.from_user.id
    session = get_state(uid)
    st = session["state"]
    data = session["data"]

    # ---- GROUP: selecting employees (toggle) ----
    if st == "group_select_employees":
        if cq.data == "empgrp_cancel":
            clear_state(uid)
            try:
                await cq.message.edit_text("❌ ثبت گروهی لغو شد.")
            except:
                pass
            await cq.answer()
            return

        if cq.data == "empgrp_confirm":
            selected = data.get("selected", [])
            if not selected:
                await cq.answer("حداقل یک کارمند انتخاب کنید!", show_alert=True)
                return
            # move to project selection
            set_state(uid, "group_project", {"employees": selected})
            try:
                await cq.message.edit_text("پروژه را انتخاب کنید:", reply_markup=projects_inline())
            except:
                await cq.message.answer("پروژه را انتخاب کنید:", reply_markup=projects_inline())
            await cq.answer()
            return

        if cq.data.startswith("empgrp_"):
            try:
                emp_id = int(cq.data.split("_")[1])
            except:
                await cq.answer()
                return
            selected = data.get("selected", [])
            if emp_id in selected:
                selected.remove(emp_id)
            else:
                selected.append(emp_id)
            data["selected"] = selected
            set_state(uid, "group_select_employees", data)
            try:
                await cq.message.edit_reply_markup(reply_markup=inline_employee_multi(selected))
            except:
                try:
                    await cq.message.edit_text("کارمندان را انتخاب کنید (✔ یعنی انتخاب شده):", reply_markup=inline_employee_multi(selected))
                except:
                    pass
            await cq.answer()
            return

    # ---- GROUP: choose project ----
    if st == "group_project":
        if cq.data == "grpprj_cancel":
            clear_state(uid)
            try:
                await cq.message.edit_text("❌ ثبت گروهی لغو شد.")
            except:
                pass
            await cq.answer()
            return
        if cq.data.startswith("grpprj_"):
            try:
                proj_id = int(cq.data.split("_")[1])
            except:
                await cq.answer()
                return
            data["project_id"] = proj_id
            # go to calendar selection
            jd = jdatetime.date.today()
            data["jdate"] = jd
            set_state(uid, "group_calendar", data)
            try:
                await cq.message.edit_text("📅 تاریخ را انتخاب کنید:", reply_markup=calendar_inline(jd))
            except:
                await cq.message.answer("📅 تاریخ را انتخاب کنید:", reply_markup=calendar_inline(jd))
            await cq.answer()
            return

    # ---- CALENDAR: could be for single or group flows ----
    if st in ("await_calendar", "group_calendar"):
        jdate = data.get("jdate", jdatetime.date.today())

        if cq.data == "cal_prev":
            jdate -= jdatetime.timedelta(days=1)
        elif cq.data == "cal_next":
            jdate += jdatetime.timedelta(days=1)
        elif cq.data == "cal_today":
            jdate = jdatetime.date.today()
        elif cq.data == "cal_yesterday":
            jdate = jdatetime.date.today() - jdatetime.timedelta(days=1)
        elif cq.data == "cal_cancel":
            clear_state(uid)
            try:
                await cq.message.edit_text("❌ ثبت لغو شد.")
            except:
                pass
            await cq.answer()
            return
        elif cq.data == "cal_confirm":
            # confirm date
            date_sh = jdate_to_str(jdate)
            if st == "await_calendar":
                data["date_jsh"] = date_sh
                set_state(uid, "await_hours", data)
                try:
                    await cq.message.edit_text(f"تاریخ انتخاب شد: {date_sh}\nلطفا ساعت کاری (مثلا 8 یا 7.5) را وارد کنید:", reply_markup=back_kb())
                except:
                    await cq.message.answer(f"تاریخ انتخاب شد: {date_sh}\nلطفا ساعت کاری (مثلا 8 یا 7.5) را وارد کنید:", reply_markup=back_kb())
                await cq.answer()
                return
            else:  # group_calendar
                data["date_jsh"] = date_sh
                set_state(uid, "group_hours", data)
                try:
                    await cq.message.edit_text(f"تاریخ گروهی انتخاب شد: {date_sh}\nلطفا ساعت کاری مشترک را وارد کنید:")
                except:
                    await cq.message.answer(f"تاریخ گروهی انتخاب شد: {date_sh}\nلطفا ساعت کاری مشترک را وارد کنید:")
                await cq.answer()
                return

        # update jdate in state and update keyboard
        data["jdate"] = jdate
        if st == "await_calendar":
            set_state(uid, "await_calendar", data)
        else:
            set_state(uid, "group_calendar", data)
        try:
            await cq.message.edit_reply_markup(reply_markup=calendar_inline(jdate))
        except:
            try:
                await cq.message.edit_text("تاریخ را انتخاب کنید:", reply_markup=calendar_inline(jdate))
            except:
                pass
        await cq.answer()
        return

    # default
    await cq.answer()

# ------------------------
# EXPORT ALL EXCEL
# ------------------------
async def generate_and_send_excel(msg: types.Message):
    rows = cursor.execute("""
        SELECT w.id, e.name, p.name, w.date, w.hours, w.overtime
        FROM work_logs w
        LEFT JOIN employees e ON e.id=w.employee_id
        LEFT JOIN projects p ON p.id=w.project_id
        ORDER BY w.id
    """).fetchall()

    if not rows:
        await msg.answer("هیچ رکوردی وجود ندارد.", reply_markup=main_menu())
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "all"
    ws.append(["log_id", "employee", "project", "date", "hours", "overtime"])
    for r in rows:
        ws.append([r[0], r[1], r[2], r[3], r[4], r[5]])

    wb.save(EXCEL_PATH)
    await msg.answer("در حال ارسال فایل اکسل...")
    await bot.send_document(msg.chat.id, types.FSInputFile(EXCEL_PATH))
    try: os.remove(EXCEL_PATH)
    except: pass

# ------------------------
# RUN BOT
# ------------------------
async def main():
    print("Bot is running...")
    await dp.start_polling(bot, timeout=30)

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except Exception as e:
        print("Bot stopped with error:", e)
